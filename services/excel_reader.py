"""Read campaign names from the Excel workbook."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from models.campaign import CampaignRecord


HEADER_LIKE_VALUES = {"campaign", "campaign name", "name"}


def read_campaign_records(excel_path: Path, sheet_name: str | None = None) -> list[CampaignRecord]:
    """Read ordered campaign names from an Excel workbook.

    Empty rows are ignored. The first non-empty cell in each row is treated as the campaign name.
    A first-row header-like value is ignored when it clearly looks like a column label.
    """

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    workbook = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        records: list[CampaignRecord] = []

        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            campaign_name = _first_non_empty_cell(row)
            if campaign_name is None:
                continue

            if row_index == 1 and campaign_name.strip().lower() in HEADER_LIKE_VALUES:
                continue

            records.append(CampaignRecord(index=len(records) + 1, name=campaign_name.strip()))

        return records
    finally:
        workbook.close()


def _first_non_empty_cell(row: Iterable[object]) -> str | None:
    for cell_value in row:
        if cell_value is None:
            continue

        text = str(cell_value).strip()
        if text:
            return text

    return None
